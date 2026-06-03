import io
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from datetime import datetime


def aplicar_formato_excel_bytes(buffer):
    wb = load_workbook(buffer)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center")
                cell.border = thin_border

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for column_cells in ws.columns:
            max_length = 0
            col = column_cells[0].column
            for cell in column_cells:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[get_column_letter(col)].width = min(max_length + 3, 80)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def generar_excel_bcp(df, total_itf, total_gastos, df_itf, df_gastos):
    df_resumen = (
        df.groupby("Concepto")
        .agg(**{
            "Cantidad de operaciones": ("Monto", "count"),
            "Monto Total": ("Monto", "sum")
        })
        .reset_index()
    )
    df_resumen["Monto Total"] = df_resumen["Monto Total"].round(2)

    df_asiento = pd.DataFrame([
        {
            "Concepto": "IMPUESTO ITF",
            "Cantidad de operaciones": len(df_itf),
            "Monto Total": round(total_itf, 2),
            "Comentario": "Registrar un solo asiento mensual por ITF"
        },
        {
            "Concepto": "GASTOS BANCARIOS",
            "Cantidad de operaciones": len(df_gastos),
            "Monto Total": round(total_gastos, 2),
            "Comentario": "Registrar un solo asiento mensual por comisiones, mantenimientos y envío de estado de cuenta"
        }
    ])

    cols_itf = ["Fecha", "Monto", "Concepto detectado", "Archivo", "Línea original"]
    cols_gastos = ["Fecha", "Monto", "Concepto detectado", "Archivo", "Línea original"]

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_resumen.to_excel(writer, sheet_name="Resumen", index=False)
        df_itf[cols_itf].to_excel(writer, sheet_name="Detalle ITF", index=False)
        df_gastos[cols_gastos].to_excel(writer, sheet_name="Detalle Gastos", index=False)
        df_asiento.to_excel(writer, sheet_name="Asiento sugerido", index=False)

    buffer.seek(0)
    return aplicar_formato_excel_bytes(buffer)


def generar_excel_interbank(df, total_itf, total_gastos, df_itf, df_gastos, columnas_ordenadas):
    df_resumen = (
        df.groupby(["Año", "Mes", "Moneda", "Símbolo", "Concepto"], as_index=False)
        .agg(**{
            "Cantidad de operaciones": ("Monto", "count"),
            "Monto Total": ("Monto", "sum")
        })
    )
    df_resumen["Monto Total"] = df_resumen["Monto Total"].round(2)

    df_asiento = pd.DataFrame([
        {
            "Concepto": "IMPUESTO ITF",
            "Cantidad de operaciones": len(df_itf),
            "Monto Total": round(total_itf, 2),
            "Comentario": "Registrar un solo asiento mensual por ITF"
        },
        {
            "Concepto": "GASTOS BANCARIOS",
            "Cantidad de operaciones": len(df_gastos),
            "Monto Total": round(total_gastos, 2),
            "Comentario": "Registrar un solo asiento mensual por comisiones, mantenimientos y envío de estado de cuenta"
        }
    ])

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_resumen.to_excel(writer, sheet_name="Resumen", index=False)
        df_itf.to_excel(writer, sheet_name="Detalle ITF", index=False)
        df_gastos.to_excel(writer, sheet_name="Detalle Gastos", index=False)
        df_asiento.to_excel(writer, sheet_name="Asiento sugerido", index=False)
        df.to_excel(writer, sheet_name="Base completa", index=False)

    buffer.seek(0)
    return aplicar_formato_excel_bytes(buffer)


def generar_excel_bbva(df, df_itf, df_gastos, columnas_ordenadas):
    df_resumen = (
        df.groupby(["Año", "Mes", "Moneda", "Símbolo", "Concepto"], as_index=False)
        .agg(**{
            "Cantidad de operaciones": ("Monto", "count"),
            "Monto Total": ("Monto", "sum")
        })
    )
    df_resumen["Monto Total"] = df_resumen["Monto Total"].round(2)

    df_asiento = df_resumen.copy()
    df_asiento["Comentario"] = df_asiento["Concepto"].apply(
        lambda x: "Registrar un solo asiento mensual por ITF"
        if x == "IMPUESTO ITF"
        else "Registrar un solo asiento mensual por comisiones y gastos bancarios"
    )

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_resumen.to_excel(writer, sheet_name="Resumen", index=False)
        df_itf.to_excel(writer, sheet_name="Detalle ITF", index=False)
        df_gastos.to_excel(writer, sheet_name="Detalle Gastos", index=False)
        df_asiento.to_excel(writer, sheet_name="Asiento sugerido", index=False)
        df.to_excel(writer, sheet_name="Base completa", index=False)

    buffer.seek(0)
    return aplicar_formato_excel_bytes(buffer)


def nombre_archivo_reporte(banco: str) -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"reporte_itf_gastos_{banco.lower()}_{ts}.xlsx"
